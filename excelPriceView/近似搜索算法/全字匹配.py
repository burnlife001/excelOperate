import tkinter as tk
from tkinter import ttk
import openpyxl

def search_keyword_in_excel(file_path, keyword):
    # 加载Excel文件
    workbook = openpyxl.load_workbook(file_path)
    
    # 遍历所有工作表
    matches = []
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        
        # 遍历所有单元格
        for row in sheet.iter_rows(values_only=True):
            for col_index, cell_value in enumerate(row, start=1):
                if cell_value and keyword.lower() in str(cell_value).lower():
                    matches.append((sheet_name, row[0], col_index, cell_value))
    return matches

def update_matches(event):
    keyword = entry.get()
    matches = search_keyword_in_excel('Iphone.xlsx', keyword)
    listbox.delete(0, tk.END)
    for match in matches:
        listbox.insert(tk.END, f"工作表 '{match[0]}' 的第 {match[1]} 行, 第 {match[2]} 列: {match[3]}")

# 创建主窗口
root = tk.Tk()
root.title("Excel Keyword Search")

# 创建搜索框
entry = tk.Entry(root)
entry.pack(pady=10)
entry.bind('<KeyRelease>', update_matches)

# 创建显示匹配结果的列表框
listbox = tk.Listbox(root, width=80, height=20)
listbox.pack(pady=10)

root.mainloop()