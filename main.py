import pandas as pd
from deep_translator import GoogleTranslator
from openpyxl import load_workbook
from copy import copy

# 宏定义
SOURCE_COLUMN = '提示词(CN)'
TARGET_COLUMN = '提示词(EN)'
SHEET_NAME = 'aicoder'

# 初始化谷歌翻译器
translator = GoogleTranslator(source='zh-CN', target='en')  # 中文 --> 英语

# 读取Excel文件
file_path = 'prompt.xlsx'
df = pd.read_excel(file_path, sheet_name=SHEET_NAME)

# 加载工作簿
wb = load_workbook(file_path)
ws = wb[SHEET_NAME]

# 翻译并写入目标列
for index, row in df.iterrows():
    prompt = row[SOURCE_COLUMN]
    if pd.notna(prompt):  # 检查是否为 NaN
        # 使用谷歌翻译引擎
        translation = translator.translate(prompt)
        # 写入目标列
        cell = ws.cell(row=index+2, column=df.columns.get_loc(TARGET_COLUMN)+1)
        cell.value = translation
        source_cell = ws.cell(row=index+2, column=df.columns.get_loc(SOURCE_COLUMN)+1)
        cell.font = copy(source_cell.font)
        cell.border = copy(source_cell.border)
        cell.fill = copy(source_cell.fill)
        cell.number_format = copy(source_cell.number_format)
        cell.protection = copy(source_cell.protection)
        cell.alignment = copy(source_cell.alignment)

# 保存修改后的Excel文件
wb.save(file_path)